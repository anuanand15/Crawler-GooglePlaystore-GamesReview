import os
import time
import openpyxl
import xlrd
from selenium import webdriver
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium.common.exceptions import NoSuchElementException


def Page():
    wb = load_workbook(filename="applist.xlsx")
    row = 1
    cwd=os.getcwd()
    os.chdir(cwd)
    chrome_driver_path = os.path.join(cwd, "chromedriver.exe")
    driver = webdriver.Chrome(executable_path=chrome_driver_path)
    driver.implicitly_wait(5)
    driver.maximize_window()
    wb1 = xlrd.open_workbook("applist.xlsx")
    sheet = wb1.sheet_by_index(0)

    for i in range(sheet.nrows - 1):
        y = sheet.cell(row, 1).value
        wb2 = Workbook()

        z = y +".xlsx"
        wb2.save(z)
        ws = wb2.create_sheet(z)
        print(y)
        url = "https://play.google.com/store/apps/details?id=" + y + "&hl=en&showAllReviews=true"
        driver.get(url)
        driver.implicitly_wait(10)
        for i in range(5):
            last_height = driver.execute_script("return document.body.scrollHeight")
            print(last_height)

            while True:
                # Scroll down to bottom
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

                # Wait to load page
                time.sleep(15)

                # Calculate new scroll height and compare with last scroll height
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height
            time.sleep(15)
            try:
                element = driver.find_element_by_css_selector(".RveJvd")
                driver.execute_script("arguments[0].scrollIntoView(true);", element)
                time.sleep(12)
                element.click()
                print("Clicked on show more")
            except NoSuchElementException:
                print("Show more button not found")

            time.sleep(15)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")


        content = driver.execute_script("return document.body.innerHTML")
        print("content taken")
        soup = BeautifulSoup(content, "html.parser")
        body = soup.find_all("span", {'jsname': "bN97Pc"})
        #body = soup.find_all("span", {'jsname': "fbQN7e"})
        print(body)

        date = soup.find_all("span", {'class': "p2TkOb"})
        r = 1
        s = 1

        for div in body:
            review = div.get_text()
            print(review )
            time.sleep(2)
            try:
                ws.cell(row=r, column=1).value = review.encode("utf-8",errors="ignore")
            except openpyxl.utils.exceptions.IllegalCharacterError:
                pass
            r = r+1

        for d in date:
            data = d.get_text()
            time.sleep(2)
            ws.cell(row=s, column=2).value = data
            s = s+1

        row=row+1
        wb.save("applist.xlsx")
        wb2.save(z)
        wb2.close()
    wb.close()


Page()


